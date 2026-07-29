"""
Analytics com Pandas.
Pandas não suporta drivers async — usa engine síncrono derivado da mesma URL.
As queries são read-only e leves, sem impacto no loop assíncrono principal.
"""

import io
import json
import logging
from datetime import datetime, timedelta, timezone

import pandas as pd
from sqlalchemy import create_engine, text

from app.core.config import get_settings
# Sem filtros de grupo ou owner — todos os setores monitorados

logger = logging.getLogger(__name__)

_DAYS_LABELS   = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
_CLOSED_STATES = ("closed", "fechado", "merged", "resolvido")

# Customers a excluir do Top Ofensores — emails de sistema/bot sem significado operacional
_OFFENDER_CUSTOMER_BLOCKLIST_DOMAINS = {
    "flux.net.br",
    "easinfo.com.br",   # emails automáticos do EasInfo
    "datora.net",       # tudo do domínio Datora
}
_OFFENDER_CUSTOMER_BLOCKLIST_EXACT = {
    "noreply@easinfo.com.br",
    "vsb-noreply@vsb.net.br",             # notificações automáticas VSB
    "portabilidade@flux.net.br",          # caixa de portabilidade
    "mailer-daemon@sh-pro122.hostgator.com.br",  # bounces automáticos
    "do_not_reply@idtexpress.com",        # notificações IDT
}

# Grupos não-operacionais — excluídos de todas as métricas
_OFFENDER_GROUP_BLOCKLIST: frozenset[str] = frozenset({
    "informativo",
    "portabilidade",
    "ativações de tn's",
    "portal ativação - número",
    "saintes",
    "portabilidade > portabilidade interna",
})

# Grupos excluídos ESPECIFICAMENTE do Top Ofensores (análise de clientes) —
# além dos não-operacionais acima, tira Saldo (mantém Saldo em Volume/Analistas).
_OFFENDER_TOPCLIENT_GROUP_EXCLUDE: frozenset[str] = _OFFENDER_GROUP_BLOCKLIST | frozenset({
    "saldo",
})

# Títulos (trechos, minúsculo) que denunciam ruído no Top Ofensores.
_OFFENDER_TITLE_BLOCKLIST_SUBSTR = (
    "portabilidade",
    "goodsales",
    "spam",
)

# Destinatários (campo "To") que indicam encaminhamento a caixas de sistema —
# tira do Top Ofensores mesmo que o remetente/título sejam de um cliente real.
_OFFENDER_RECIPIENT_BLOCKLIST_SUBSTR = (
    "portabilidade@flux.net.br",
)

# Analistas que atuam exclusivamente em grupos não-operacionais
_ANALYST_BLOCKLIST: frozenset[str] = frozenset({
    "gabriel.alves@flux.net.br",
    "hanna.avila@flux.net.br",
    "jessica.cerveira@flux.net.br",
    "ryan.santos@flux.net.br",       # exibido como "Ryan Chaves" no Zammad
    "carlos@flux.net.br",            # inativo
    "carlos.roberto@flux.net.br",    # inativo
    "infrarotas@flux.net.br",        # inativo
    "thiagoventer@gmail.com",        # duplicado — usar thiago.venter.ext@flux.net.br
})

# Fragmentos SQL reutilizáveis (compatível com SQLite — sem parâmetros bind para listas)
# Apostrofos dentro dos nomes são escapados dobrando-os (padrão SQL: '' = literal ')
def _sql_str(s: str) -> str:
    return "'" + s.replace("'", "''") + "'"

_GROUP_SQL_LIST   = ", ".join(_sql_str(g) for g in _OFFENDER_GROUP_BLOCKLIST)
_ANALYST_SQL_LIST = ", ".join(_sql_str(a) for a in _ANALYST_BLOCKLIST)
_SQL_NOT_EXCLUDED_GROUP   = f'LOWER(TRIM(COALESCE("group", \'\'))) NOT IN ({_GROUP_SQL_LIST})'
_SQL_NOT_EXCLUDED_ANALYST = f"LOWER(TRIM(COALESCE(owner, ''))) NOT IN ({_ANALYST_SQL_LIST})"


def _sync_engine():
    url = get_settings().database_url.replace("+aiosqlite", "")
    return create_engine(url, connect_args={"check_same_thread": False})


def _fmt(seconds: float) -> str:
    s = int(seconds)
    if s < 60:   return f"{s}s"
    if s < 3600: return f"{s // 60}m"
    if s < 86400:
        h, m = divmod(s, 3600)
        return f"{h}h {m // 60}m" if m else f"{h}h"
    d, rem = divmod(s, 86400)
    h = rem // 3600
    return f"{d}d {h}h" if h else f"{d}d"


def _resolve_dates(start_date, end_date, default_days=30):
    today = datetime.now(timezone.utc).date()
    if not end_date:   end_date   = today.isoformat()
    if not start_date: start_date = (today - timedelta(days=default_days - 1)).isoformat()
    return start_date, end_date


def _load_tickets(start_date: str, end_date: str, engine) -> pd.DataFrame:
    today   = datetime.now(timezone.utc).date().isoformat()
    start_s = f"{start_date} 00:00:00"
    end_s   = (
        datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        if end_date >= today
        else f"{end_date} 23:59:59"
    )
    query = text(
        "SELECT * FROM tickets"
        " WHERE datetime(created_at) >= datetime(:start)"
        "   AND datetime(created_at) <= datetime(:end)"
    )
    df = pd.read_sql(query, con=engine, params={"start": start_s, "end": end_s})
    if not df.empty:
        df["created_at"] = pd.to_datetime(df["created_at"], utc=True, errors="coerce")
    return df


# ── Top Ofensores ─────────────────────────────────────────────────────────────

def compute_top_offenders(by="customer", start_date=None, end_date=None, limit=5) -> dict:
    if by not in ("customer", "group"):
        by = "customer"
    start_date, end_date = _resolve_dates(start_date, end_date)
    engine = _sync_engine()
    df = _load_tickets(start_date, end_date, engine)
    if df.empty:
        return {"by": by, "start_date": start_date, "end_date": end_date, "items": []}

    # Conta TODOS os tickets criados no período (independente do estado atual).
    # O filtro de "somente ativos" era incorreto para meses passados — a maioria
    # já está fechada, fazendo os contadores históricos aparecerem zerados.
    period_df = df.copy()
    col = by
    if col not in period_df.columns or period_df[col].isna().all():
        return {"by": by, "start_date": start_date, "end_date": end_date, "items": []}
    period_df[col] = period_df[col].fillna("(sem nome)").str.strip()

    if by == "customer":
        def _is_blocked_customer(v: str) -> bool:
            v_lower = v.lower()
            if v_lower in _OFFENDER_CUSTOMER_BLOCKLIST_EXACT:
                return True
            if "@" in v_lower:
                domain = v_lower.split("@", 1)[1]
                if domain in _OFFENDER_CUSTOMER_BLOCKLIST_DOMAINS:
                    return True
            return False
        period_df = period_df[~period_df[col].apply(_is_blocked_customer)]

        # Exclui tickets de grupos não-operacionais / Saldo (ruído na análise de clientes)
        if "group" in period_df.columns:
            grp = period_df["group"].fillna("").str.lower().str.strip()
            period_df = period_df[~grp.isin(_OFFENDER_TOPCLIENT_GROUP_EXCLUDE)]

        # Exclui tickets cujo título contém trechos de ruído (portabilidade, spam, goodsales)
        if "title" in period_df.columns and not period_df.empty:
            title_l = period_df["title"].fillna("").str.lower()
            mask_title = title_l.apply(
                lambda t: any(sub in t for sub in _OFFENDER_TITLE_BLOCKLIST_SUBSTR)
            )
            period_df = period_df[~mask_title]

        # Exclui tickets ENCAMINHADOS para caixas de sistema (destinatário/"To")
        if "recipient" in period_df.columns and not period_df.empty:
            recip_l = period_df["recipient"].fillna("").str.lower()
            mask_recip = recip_l.apply(
                lambda r: any(sub in r for sub in _OFFENDER_RECIPIENT_BLOCKLIST_SUBSTR)
            )
            period_df = period_df[~mask_recip]
    else:
        period_df = period_df[~period_df[col].str.lower().str.strip().isin(_OFFENDER_GROUP_BLOCKLIST)]

    if period_df.empty:
        return {"by": by, "start_date": start_date, "end_date": end_date, "items": []}

    top_names = period_df.groupby(col).size().nlargest(limit).index.tolist()
    items = []
    for name in top_names:
        subset = period_df[period_df[col] == name]
        items.append({
            "name": name,
            "ticket_count": len(subset),
            "priority_breakdown": subset.groupby("priority").size().to_dict(),
        })
    return {"by": by, "start_date": start_date, "end_date": end_date, "items": items}


# ── Heatmap ───────────────────────────────────────────────────────────────────

def compute_heatmap(start_date=None, end_date=None) -> dict:
    start_date, end_date = _resolve_dates(start_date, end_date, default_days=90)
    engine = _sync_engine()
    df = _load_tickets(start_date, end_date, engine)
    if df.empty:
        return {"start_date": start_date, "end_date": end_date, "max_count": 0,
                "days_labels": _DAYS_LABELS, "data": []}
    df = df.dropna(subset=["created_at"])
    df["day"]  = df["created_at"].dt.dayofweek
    df["hour"] = df["created_at"].dt.hour
    heat = df.groupby(["day", "hour"]).size().reset_index(name="count")
    full_index = pd.MultiIndex.from_product([range(7), range(24)], names=["day", "hour"])
    heat = (heat.set_index(["day", "hour"]).reindex(full_index, fill_value=0).reset_index())
    max_count = int(heat["count"].max())
    return {"start_date": start_date, "end_date": end_date, "max_count": max_count,
            "days_labels": _DAYS_LABELS,
            "data": heat[["day", "hour", "count"]].to_dict(orient="records")}


# ── FRT ───────────────────────────────────────────────────────────────────────

def _frt_status(mean_seconds: float) -> str:
    if mean_seconds < 1_800: return "good"
    if mean_seconds < 3_600: return "warning"
    return "critical"


def _empty_frt(days: int) -> dict:
    return {"count": 0, "period_days": days,
            "mean_seconds": 0.0, "mean_formatted": "—",
            "p50_seconds": 0.0,  "p50_formatted": "—",
            "p90_seconds": 0.0,  "p90_formatted": "—",
            "p95_seconds": 0.0,  "p95_formatted": "—"}


def compute_frt_stats(days=30) -> dict:
    start_date, end_date = _resolve_dates(None, None, default_days=days)
    engine = _sync_engine()
    df = _load_tickets(start_date, end_date, engine)
    if df.empty or "frt_seconds" not in df.columns:
        return _empty_frt(days)
    frt = df["frt_seconds"].dropna().astype(float)
    if frt.empty:
        return _empty_frt(days)
    mean_s = frt.mean()
    return {"count": len(frt), "period_days": days,
            "mean_seconds": round(mean_s, 1), "mean_formatted": _fmt(mean_s),
            "p50_seconds": round(frt.quantile(0.50), 1), "p50_formatted": _fmt(frt.quantile(0.50)),
            "p90_seconds": round(frt.quantile(0.90), 1), "p90_formatted": _fmt(frt.quantile(0.90)),
            "p95_seconds": round(frt.quantile(0.95), 1), "p95_formatted": _fmt(frt.quantile(0.95))}


def compute_frt_today() -> dict:
    engine = _sync_engine()
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    query = text("SELECT frt_seconds FROM tickets WHERE created_at >= :since AND frt_seconds IS NOT NULL")
    df = pd.read_sql(query, con=engine, params={"since": today_start.isoformat()})
    date_str = today_start.date().isoformat()
    if df.empty:
        return {"date": date_str, "count": 0,
                "mean_seconds": 0.0, "mean_formatted": "—",
                "p50_seconds": 0.0,  "p50_formatted": "—",
                "p95_seconds": 0.0,  "p95_formatted": "—",
                "status": "good"}
    frt    = df["frt_seconds"].dropna().astype(float)
    mean_s = frt.mean()
    return {"date": date_str, "count": len(frt),
            "mean_seconds": round(mean_s, 1), "mean_formatted": _fmt(mean_s),
            "p50_seconds": round(frt.quantile(0.50), 1), "p50_formatted": _fmt(frt.quantile(0.50)),
            "p95_seconds": round(frt.quantile(0.95), 1), "p95_formatted": _fmt(frt.quantile(0.95)),
            "status": _frt_status(mean_s)}


# ── Carga por Analista ────────────────────────────────────────────────────────

def _month_to_dates(month: str):
    """'2026-05' → ('2026-05-01', '2026-05-31')"""
    import calendar
    y, m = int(month[:4]), int(month[5:7])
    last = calendar.monthrange(y, m)[1]
    return f"{y:04d}-{m:02d}-01", f"{y:04d}-{m:02d}-{last:02d}"


def compute_analyst_load(start_date=None, end_date=None, month=None) -> dict:
    """Carga histórica por analista.
    Suporta:
      • month='YYYY-MM'  → cobre o mês inteiro
      • start_date + end_date → intervalo livre
      • sem parâmetros  → todos os tickets ativos (sem filtro de data)
    """
    engine = _sync_engine()
    base_filter = (
        " WHERE owner IS NOT NULL AND owner != '' AND owner != '-'"
        " AND LOWER(COALESCE(owner, '')) NOT LIKE 'auto-%'"
        " AND (owner_id IS NULL OR owner_id != 1)"
        f" AND {_SQL_NOT_EXCLUDED_GROUP}"
        f" AND {_SQL_NOT_EXCLUDED_ANALYST}"
    )

    today = datetime.now(timezone.utc).date().isoformat()

    if month:
        start_date, end_date = _month_to_dates(month)

    if start_date and end_date:
        start_s = f"{start_date} 00:00:00"
        end_s   = (
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
            if end_date >= today else f"{end_date} 23:59:59"
        )
        query = text(
            "SELECT owner FROM tickets" + base_filter +
            " AND datetime(created_at) >= datetime(:start)"
            " AND datetime(created_at) <= datetime(:end)"
        )
        df = pd.read_sql(query, con=engine, params={"start": start_s, "end": end_s})
    else:
        df = pd.read_sql(text("SELECT owner FROM tickets" + base_filter), con=engine)

    if df.empty:
        return {"items": []}

    df["owner"] = df["owner"].fillna("(sem dono)").str.strip()
    df = df[df["owner"] != ""]
    ranked = (df.groupby("owner").size()
                .reset_index(name="ticket_count")
                .sort_values("ticket_count", ascending=False))
    return {"items": ranked.to_dict(orient="records")}


def compute_analyst_day_detail(date: str) -> dict:
    """Tickets ATIVOS (exceto resolvido/fechado) por analista.

    Para qualquer data selecionada retorna todos os tickets criados até
    aquela data que ainda NÃO foram resolvidos/fechados — ou seja, a
    carga atual de cada analista vista a partir daquele ponto.
    """
    engine = _sync_engine()
    today = datetime.now(timezone.utc).date().isoformat()
    end_s = (
        datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        if date >= today else f"{date} 23:59:59"
    )
    active_states = "'em atendimento', 'aguardando cliente', 'aguardando terceiros', 'escalonado dev', 'escalonado rotas', 'escalonado infra'"

    query = text(
        f"SELECT owner, number, title, state FROM tickets"
        f" WHERE LOWER(TRIM(state)) IN ({active_states})"
        f"   AND datetime(created_at) <= datetime(:end)"
        f"   AND owner IS NOT NULL AND owner != '' AND owner != '-'"
        f"   AND LOWER(COALESCE(owner, '')) NOT LIKE 'auto-%'"
        f"   AND (owner_id IS NULL OR owner_id != 1)"
        f"   AND {_SQL_NOT_EXCLUDED_GROUP}"
        f"   AND {_SQL_NOT_EXCLUDED_ANALYST}"
    )
    df = pd.read_sql(query, con=engine, params={"end": end_s})
    if df.empty:
        return {"date": date, "analysts": [], "total": 0}

    df["owner"] = df["owner"].str.strip()
    analysts = []
    for owner, grp in df.groupby("owner", sort=False):
        grp_sorted = grp.sort_values("number", ascending=False)
        analysts.append({
            "owner": owner,
            "ticket_count": len(grp_sorted),
            "tickets": grp_sorted[["number", "title", "state"]].to_dict(orient="records"),
        })
    analysts.sort(key=lambda x: x["ticket_count"], reverse=True)
    return {"date": date, "analysts": analysts, "total": len(df)}


def compute_daily_volume(month: str | None = None, owner: str | None = None) -> dict:
    """
    Sem owner → visão global (criados / resolvidos / fechados + snapshots de estado).
    Com owner  → visão analista (ativos no dia vs fechados no dia).
    """
    engine = _sync_engine()

    today_iso = datetime.now(timezone.utc).date().isoformat()
    if month:
        start_date, end_date = _month_to_dates(month)
        # Nunca exibir dias futuros — limita ao dia atual
        if end_date > today_iso:
            end_date = today_iso
    else:
        end_date   = today_iso
        start_date = (datetime.now(timezone.utc).date() - timedelta(days=29)).isoformat()

    start_s = f"{start_date} 00:00:00"
    end_s   = (
        datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        if end_date >= today_iso else f"{end_date} 23:59:59"
    )

    all_days = pd.date_range(start=start_date, end=end_date, freq="D")

    # ── Visão por analista ────────────────────────────────────────────────────
    if owner:
        # Carrega tickets do analista que tiveram atividade no período:
        # updated_at no período OU close_at no período
        q = text(
            "SELECT updated_at, close_at, state FROM tickets"
            " WHERE LOWER(TRIM(owner)) = LOWER(TRIM(:owner))"
            "   AND ("
            "     (datetime(updated_at) >= datetime(:start) AND datetime(updated_at) <= datetime(:end))"
            "     OR (close_at IS NOT NULL AND datetime(close_at) >= datetime(:start) AND datetime(close_at) <= datetime(:end))"
            "   )"
        )
        df_t = pd.read_sql(q, con=engine, params={"owner": owner, "start": start_s, "end": end_s})
        df_t["updated_dt"] = pd.to_datetime(df_t["updated_at"], errors="coerce")
        df_t["close_dt"]   = pd.to_datetime(df_t["close_at"],   errors="coerce")
        df_t["state_lc"]   = df_t["state"].str.lower().str.strip()

        _CONCLUIDO = {"resolvido", "closed", "fechado", "merged"}
        points = []
        for day in all_days:
            day_date   = day.date()
            is_weekend = day_date.weekday() >= 5  # 5=Sábado, 6=Domingo

            # Fim de semana: Zammad auto-resolve tickets "Aguardando Cliente" sem
            # ação do analista — zeramos ambos os contadores nesses dias.
            if is_weekend:
                trabalhados = 0
                resolvidos  = 0
            else:
                # Resolvidos: close_at neste dia (dia útil) em estado conclusivo
                resolvidos_mask = (
                    df_t["close_dt"].notna() &
                    (df_t["close_dt"].dt.date == day_date) &
                    df_t["state_lc"].isin(_CONCLUIDO)
                )
                resolvidos = int(resolvidos_mask.sum())

                # Atendidos: updated_at neste dia, excluindo já contados como resolvidos
                trabalhados = int((
                    (df_t["updated_dt"].dt.date == day_date) &
                    ~resolvidos_mask
                ).sum())

            points.append({
                "date":        day.strftime("%d/%m"),
                "dia":         day.strftime("%Y-%m-%d"),
                "trabalhados": trabalhados,
                "resolvidos":  resolvidos,
            })

        return {
            "mode":       "analyst",
            "month":      month or "custom",
            "start_date": start_date,
            "end_date":   end_date,
            "owner":      owner,
            "has_snap":   False,
            "points":     points,
        }

    # ── Visão global ──────────────────────────────────────────────────────────
    params_base = {"start": start_s, "end": end_s}

    q_criados = text(
        f"SELECT DATE(created_at) as dia, COUNT(*) as criados FROM tickets"
        f" WHERE datetime(created_at) >= datetime(:start)"
        f"   AND datetime(created_at) <= datetime(:end)"
        f"   AND {_SQL_NOT_EXCLUDED_GROUP}"
        f" GROUP BY dia ORDER BY dia"
    )
    q_fechados = text(
        f"SELECT DATE(close_at) as dia,"
        f"  SUM(CASE WHEN LOWER(TRIM(state)) = 'resolvido' THEN 1 ELSE 0 END) as resolvidos,"
        f"  SUM(CASE WHEN LOWER(TRIM(state)) IN ('fechado','closed','merged') THEN 1 ELSE 0 END) as fechados"
        f" FROM tickets"
        f" WHERE close_at IS NOT NULL"
        f"   AND datetime(close_at) >= datetime(:start)"
        f"   AND datetime(close_at) <= datetime(:end)"
        f"   AND {_SQL_NOT_EXCLUDED_GROUP}"
        f" GROUP BY dia ORDER BY dia"
    )

    df_c = pd.read_sql(q_criados,  con=engine, params=dict(params_base))
    df_f = pd.read_sql(q_fechados, con=engine, params=dict(params_base))

    # Tickets ativos: carrega estado para calcular breakdown por estado em todos os meses
    # Exclui 'merged': tickets mesclados ficam com close_at=NULL mas não estão ativos
    _FILA_LC      = {"em atendimento", "escalonado dev", "escalonado rotas", "escalonado infra"}
    _AG_CLI_LC    = {"aguardando cliente"}
    _AG_TER_LC    = {"aguardando terceiros"}

    q_active = text(
        f"SELECT created_at, close_at, state FROM tickets"
        f" WHERE datetime(created_at) <= datetime(:end)"
        f"   AND (close_at IS NULL OR datetime(close_at) >= datetime(:start))"
        f"   AND LOWER(TRIM(state)) != 'merged'"
        f"   AND {_SQL_NOT_EXCLUDED_GROUP}"
    )
    df_act = pd.read_sql(q_active, con=engine, params=dict(params_base))
    df_act["created_dt"] = pd.to_datetime(df_act["created_at"], errors="coerce")
    df_act["close_dt"]   = pd.to_datetime(df_act["close_at"],   errors="coerce")
    df_act["state_lc"]   = df_act["state"].str.lower().str.strip()

    df = pd.DataFrame({"dia": all_days.strftime("%Y-%m-%d")})
    df = df.merge(df_c, on="dia", how="left") if not df_c.empty else df.assign(criados=0)
    df = df.merge(df_f, on="dia", how="left") if not df_f.empty else df.assign(resolvidos=0, fechados=0)
    df = df.fillna(0)
    for col in ("criados", "resolvidos", "fechados"):
        df[col] = df[col].astype(int) if col in df.columns else 0

    # Snapshots de estado — mais precisos para mai/2026+ (usados como override)
    snap_q = text(
        "SELECT key, value FROM sync_meta"
        " WHERE key LIKE 'snap_%' AND key >= :kmin AND key <= :kmax ORDER BY key"
    )
    snap_raw = pd.read_sql(snap_q, con=engine,
                           params={"kmin": f"snap_{start_date}", "kmax": f"snap_{end_date}"})
    snap_dict: dict = {}
    for _, srow in snap_raw.iterrows():
        snap = json.loads(srow["value"])
        snap_dict[srow["key"][5:]] = {
            "em_atendimento": snap.get("em_atendimento",       0),
            "ag_cliente":     snap.get("aguardando_cliente",   0),
            "ag_terceiros":   snap.get("aguardando_terceiros", 0),
        }

    has_snap = len(snap_dict) > 0
    df["date"] = pd.to_datetime(df["dia"]).dt.strftime("%d/%m")

    points = []
    for _, row in df.iterrows():
        # Linhas de estado (em_atendimento, ag_cliente, ag_terceiros):
        # Só são confiáveis quando existe snapshot salvo pelo dashboard naquele dia.
        # Para dias sem snapshot (meses anteriores ao início do dashboard), retornamos
        # None — o frontend exibe uma lacuna no gráfico em vez de dado impreciso.
        if row["dia"] in snap_dict:
            s        = snap_dict[row["dia"]]
            em_atend = s["em_atendimento"]
            ag_cli   = s["ag_cliente"]
            ag_ter   = s["ag_terceiros"]
        else:
            em_atend = None
            ag_cli   = None
            ag_ter   = None

        pt = {
            "date":           row["date"],
            "dia":            row["dia"],
            "criados":        int(row.get("criados",    0) or 0),
            "resolvidos":     int(row.get("resolvidos", 0) or 0),
            "fechados":       int(row.get("fechados",   0) or 0),
            "em_atendimento": em_atend,
            "ag_cliente":     ag_cli,
            "ag_terceiros":   ag_ter,
        }
        points.append(pt)

    return {
        "mode":       "global",
        "month":      month or "custom",
        "start_date": start_date,
        "end_date":   end_date,
        "owner":      None,
        "has_snap":   has_snap,
        "points":     points,
    }


def compute_top_offender_detail(by: str, name: str, start_date=None, end_date=None) -> dict:
    """Lista de tickets criados no período para um cliente ou grupo específico."""
    if by not in ("customer", "group"):
        by = "customer"
    start_date, end_date = _resolve_dates(start_date, end_date)
    engine  = _sync_engine()
    today   = datetime.now(timezone.utc).date().isoformat()
    start_s = f"{start_date} 00:00:00"
    end_s   = (
        datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        if end_date >= today else f"{end_date} 23:59:59"
    )
    col = "customer" if by == "customer" else '"group"'
    # Exibe TODOS os tickets criados no período (sem filtrar por estado atual).
    # O contador principal também conta todos — manter consistência.
    # Para o mês atual, prioriza tickets ainda em aberto (aparecem primeiro).
    closed = ", ".join(f"'{s}'" for s in _CLOSED_STATES)
    is_current = end_date >= today
    order = (
        f"CASE WHEN LOWER(state) NOT IN ({closed}) THEN 0 ELSE 1 END, "
        f"CAST(number AS INTEGER) DESC"
    ) if is_current else "CAST(number AS INTEGER) DESC"
    query = text(
        f"SELECT number, title, state, priority, owner FROM tickets"
        f" WHERE {col} = :name"
        f"   AND datetime(created_at) >= datetime(:start)"
        f"   AND datetime(created_at) <= datetime(:end)"
        f" ORDER BY {order}"
    )
    df = pd.read_sql(query, con=engine, params={"name": name, "start": start_s, "end": end_s})
    if df.empty:
        return {"by": by, "name": name, "start_date": start_date,
                "end_date": end_date, "tickets": [], "total": 0}

    df["owner"] = df["owner"].fillna("").str.strip()
    df["priority"] = df["priority"].fillna("").str.strip()
    tickets = df[["number", "title", "state", "priority", "owner"]].to_dict(orient="records")
    return {"by": by, "name": name, "start_date": start_date,
            "end_date": end_date, "tickets": tickets, "total": len(tickets)}


# ── Exportação Excel ──────────────────────────────────────────────────────────

def _strip_tz(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            try:
                if df[col].dt.tz is not None:
                    df[col] = df[col].dt.tz_localize(None)
            except Exception:
                pass
    return df


def generate_excel_report(start_date=None, end_date=None) -> bytes:
    start_date, end_date = _resolve_dates(start_date, end_date, default_days=30)
    today   = datetime.now(timezone.utc).date().isoformat()
    start_s = f"{start_date} 00:00:00"
    end_s   = (
        datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        if end_date >= today else f"{end_date} 23:59:59"
    )
    engine = _sync_engine()
    now    = datetime.now(timezone.utc)
    query  = text(
        "SELECT * FROM tickets"
        " WHERE datetime(created_at) >= datetime(:start)"
        "   AND datetime(created_at) <= datetime(:end)"
    )
    df = pd.read_sql(query, con=engine, params={"start": start_s, "end": end_s})
    if not df.empty:
        for col in ("created_at", "close_at", "updated_at"):
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], utc=True, errors="coerce")
        df = _strip_tz(df)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        _sheet_resumo(writer, df, now, start_date, end_date)
        _sheet_volume_diario(writer, df)
        _sheet_top_ofensores(writer, df)
        _sheet_frt(writer, df)
        _style_workbook(writer.book)
    buf.seek(0)
    return buf.read()


def _sheet_resumo(writer, df, now, start_date, end_date):
    active  = df[~df["state"].str.lower().isin(_CLOSED_STATES)] if not df.empty else df
    closed  = df[df["close_at"].notna()] if not df.empty and "close_at" in df.columns else pd.DataFrame()
    frt     = df["frt_seconds"].dropna().astype(float) if not df.empty and "frt_seconds" in df.columns else pd.Series(dtype=float)
    rows = [
        ("Período", f"{start_date} — {end_date}"),
        ("Gerado em", now.strftime("%d/%m/%Y %H:%M UTC")),
        ("", ""),
        ("── VOLUMETRIA ──", ""),
        ("Total criados no período", len(df)),
        ("Fechados no período", len(closed)),
        ("Em aberto / pendente", len(active)),
        ("", ""),
        ("── FRT (Primeira Resposta) ──", ""),
        ("Chamados com FRT medido", len(frt)),
        ("FRT Médio",  _fmt(frt.mean())         if not frt.empty else "—"),
        ("FRT P50",    _fmt(frt.quantile(0.50)) if not frt.empty else "—"),
        ("FRT P90",    _fmt(frt.quantile(0.90)) if not frt.empty else "—"),
        ("FRT P95",    _fmt(frt.quantile(0.95)) if not frt.empty else "—"),
    ]
    pd.DataFrame(rows, columns=["Indicador", "Valor"]).to_excel(writer, sheet_name="Resumo", index=False)


def _sheet_volume_diario(writer, df):
    empty = pd.DataFrame(columns=["Data", "Criados", "Fechados", "Líquido"])
    if df.empty or "created_at" not in df.columns:
        empty.to_excel(writer, sheet_name="Volume Diário", index=False)
        return
    df = df.dropna(subset=["created_at"]).copy()
    df["data"] = df["created_at"].dt.strftime("%d/%m/%Y")
    criados  = df.groupby("data").size().reset_index(name="Criados")
    fechados = pd.DataFrame(columns=["data", "Fechados"])
    if "close_at" in df.columns:
        df_c = df.dropna(subset=["close_at"]).copy()
        if not df_c.empty:
            df_c["data"] = df_c["close_at"].dt.strftime("%d/%m/%Y")
            fechados = df_c.groupby("data").size().reset_index(name="Fechados")
    vol = criados.merge(fechados, on="data", how="left").fillna(0)
    vol["Fechados"] = vol["Fechados"].astype(int)
    vol["Líquido"]  = vol["Criados"] - vol["Fechados"]
    vol.rename(columns={"data": "Data"}).to_excel(writer, sheet_name="Volume Diário", index=False)


def _sheet_top_ofensores(writer, df):
    empty = pd.DataFrame(columns=["#", "Cliente", "Chamados", "% do Total"])
    if df.empty or "customer" not in df.columns:
        empty.to_excel(writer, sheet_name="Top Ofensores", index=False)
        return
    active = df[~df["state"].str.lower().isin(_CLOSED_STATES)].copy()
    active["customer"] = active["customer"].fillna("(sem nome)")
    top = (active.groupby("customer").size().reset_index(name="Chamados")
           .sort_values("Chamados", ascending=False).head(10))
    total = top["Chamados"].sum()
    top["% do Total"] = (top["Chamados"] / max(total, 1) * 100).round(1).astype(str) + "%"
    top.insert(0, "#", range(1, len(top) + 1))
    top.rename(columns={"customer": "Cliente"}).to_excel(writer, sheet_name="Top Ofensores", index=False)


def _sheet_frt(writer, df):
    cols_needed = {"number", "priority", "group", "created_at", "frt_seconds"}
    empty = pd.DataFrame(columns=["Ticket", "Prioridade", "Setor", "Criado em", "FRT (s)", "FRT"])
    if df.empty or not cols_needed.issubset(df.columns):
        empty.to_excel(writer, sheet_name="FRT", index=False)
        return
    frt_df = df.dropna(subset=["frt_seconds"])[list(cols_needed)].copy()
    if frt_df.empty:
        empty.to_excel(writer, sheet_name="FRT", index=False)
        return
    frt_df["FRT"]       = frt_df["frt_seconds"].apply(_fmt)
    frt_df["Criado em"] = frt_df["created_at"].dt.strftime("%d/%m/%Y %H:%M")
    (frt_df.rename(columns={"number": "Ticket", "priority": "Prioridade",
                             "group": "Setor", "frt_seconds": "FRT (s)"})
     [["Ticket", "Prioridade", "Setor", "Criado em", "FRT (s)", "FRT"]]
     .sort_values("FRT (s)", ascending=False)
     .to_excel(writer, sheet_name="FRT", index=False))


def _style_workbook(wb) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    FILL  = PatternFill("solid", fgColor="0F1729")
    FONT  = Font(bold=True, color="60A5FA", size=10)
    ALIGN = Alignment(horizontal="center", vertical="center")
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = FONT; cell.fill = FILL; cell.alignment = ALIGN
        ws.row_dimensions[1].height = 18
        for col in ws.columns:
            width = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(width + 4, 48)
